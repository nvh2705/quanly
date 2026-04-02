import io

from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from flask import send_file
import os

app = Flask(__name__)
app.config['SECRET_KEY'] = 'security_123'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///nhaan.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_message = "Bạn cần đăng nhập để truy cập trang này!"
login_manager.login_message_category = "danger"
login_manager.login_view = 'login'

# --- MODELS ---
class AdminLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    content = db.Column(db.String(500), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    is_read = db.Column(db.Boolean, default=False)


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(100), nullable=False)
    fullname = db.Column(db.String(100))
    role = db.Column(db.String(20), nullable=False) # admin, nhaan, cbgv, loppho, hocvien
    unit = db.Column(db.String(50)) # Lớp hoặc Khoa
    is_read = db.Column(db.Boolean, default=False)

class MealRegistration(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    sang = db.Column(db.Boolean, default=False)
    trua = db.Column(db.Boolean, default=False)
    toi = db.Column(db.Boolean, default=False)
    user_rel = db.relationship('User', backref='meals')

class LoginHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer)
    username = db.Column(db.String(50))
    login_time = db.Column(db.DateTime)
    logout_time = db.Column(db.DateTime)
    ip_address = db.Column(db.String(50))
    content = db.Column(db.String(500))  # Thêm trường content để lưu thông tin đăng ký ăn

class ClassModel(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ten_lop = db.Column(db.String(50), unique=True, nullable=False)

class StudentStatistic(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ten = db.Column(db.String(100), nullable=False)
    lop_id = db.Column(db.Integer, db.ForeignKey('class_model.id'))
    vang_sang = db.Column(db.Integer, default=0)
    vang_trua = db.Column(db.Integer, default=0)
    vang_toi = db.Column(db.Integer, default=0)
    tien_thua = db.Column(db.Integer, default=0)
    lop_rel = db.relationship('ClassModel', backref='students')

class Message(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    sender = db.Column(db.String(50))  # username
    role = db.Column(db.String(20))    # admin / hocvien / cbgv
    content = db.Column(db.Text)
    reply_to = db.Column(db.Integer, nullable=True)
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class StaffStatistic(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    don_vi = db.Column(db.String(50))  # Khoa / Phòng
    so_luong = db.Column(db.Integer, default=0)
    vang_sang = db.Column(db.Integer, default=0)
    vang_trua = db.Column(db.Integer, default=0)
    vang_toi = db.Column(db.Integer, default=0)

# --- LOGIC HỖ TRỢ ---
def init_default_meals(user_id, start_date):
    """Tự động đăng ký T2-T6 cho học viên mới/tuần mới"""
    for i in range(5):  # Thứ 2 đến Thứ 6
        current_date = start_date + timedelta(days=i)
        check_exist = MealRegistration.query.filter_by(user_id=user_id, date=current_date).first()
        if not check_exist:
            new_meal = MealRegistration(
                user_id=user_id, 
                date=current_date, 
                sang=True, trua=True, toi=True
            )
            db.session.add(new_meal)
    db.session.commit()

@app.route("/signup", methods=["GET", "POST"])
def signup():

    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        fullname = request.form["fullname"]
        role = request.form["role"]
        unit = request.form["unit"]

        # check user tồn tại
        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            flash("Tên đăng nhập đã tồn tại!", "danger")
            return redirect(url_for("signup"))

        # hash password
        hashed_password = generate_password_hash(password)

        # tạo user mới
        new_user = User(username=username, password=hashed_password, fullname=fullname, role=role, unit=unit)
        db.session.add(new_user)
        db.session.commit()

        # Nếu là học viên, khởi tạo đăng ký ăn mặc định
        if role == 'hocvien':
            today = datetime.now().date()
            start_of_week = today - timedelta(days=today.weekday())  # Monday
            init_default_meals(new_user.id, start_of_week)

        flash("Đăng ký thành công! Vui lòng đăng nhập.", "success")
        return redirect(url_for("login"))

    return render_template("signup.html")

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# --- ROUTES ---
# 1. Khai báo kho dữ liệu thực đơn (Menu)
MENU_STORAGE = {
    'Thứ 2': {
        'sang': 'Xôi thịt kho',
        'trua': 'Thịt kho tàu, Đậu rán sốt cà, Canh cải cúc',
        'toi': 'Cá kho tộ, Rau muống luộc, Trứng đúc thịt'
    },
    'Thứ 3': {
        'sang': 'Mì tôm trứng',
        'trua': 'Gà rang gừng, Su su xào tỏi, Canh bí xanh',
        'toi': 'Thịt băm rim, Trứng rán, Bắp cải xào'
    },
    'Thứ 4': {
        'sang': 'Bánh mì, sữa mi lô',
        'trua': 'Sườn xào chua ngọt, Đậu phụ trắng, Canh rau ngót',
        'toi': 'Tôm rang thịt ba chỉ, Đỗ quả xào, Canh chua'
    },
    'Thứ 5': {
        'sang': 'Mì tôm trứng vịt lộn',
        'trua': 'Thịt chân giò luộc, Cá chiên xù, Canh mồng tơi',
        'toi': 'Bò xào cần tây, Chả lá lốt, Rau dền luộc'
    },
    'Thứ 6': {
        'sang': 'Bánh bao',
        'trua': 'Vịt quay, Đậu sốt, Canh khoai tây',
        'toi': 'Cá thu sốt cà, Rau cải chíp xào, Thịt kho'
    },
    'Thứ 7': {
        'sang': 'Bánh mì, sữa tươi',
        'trua': 'Thịt kho trứng, Bầu xào tôm',
        'toi': 'Gà luộc lá chanh, Canh miến nấu lòng gà'
    },
    'Chủ Nhật': {
        'sang': 'Mì tôm trứng',
        'trua': 'Cơm liên hoan cuối tuần',
        'toi': 'Cơm rang thập cẩm, Salad'
    }
}

# 2. Hàm bổ trợ để lấy thực đơn đúng ngày hiện tại
def get_current_menu():
    import datetime
    # Lấy thứ trong tuần (0=Thứ 2, ..., 6=Chủ Nhật)
    weekday_index = datetime.datetime.now().weekday()
    weekday_map = {0: 'Thứ 2', 1: 'Thứ 3', 2: 'Thứ 4', 3: 'Thứ 5', 4: 'Thứ 6', 5: 'Thứ 7', 6: 'Chủ Nhật'}
    
    today_name = weekday_map.get(weekday_index, 'Thứ 2')
    menu_today = MENU_STORAGE.get(today_name)
    
    return menu_today, today_name

# Định nghĩa đơn giá cố định
GIA_SANG = 15000
GIA_TRUA = 30000
GIA_TOI = 30000

def get_current_menu_with_price():
    import datetime
    weekday_index = datetime.datetime.now().weekday()
    weekday_map = {0: 'Thứ 2', 1: 'Thứ 3', 2: 'Thứ 4', 3: 'Thứ 5', 4: 'Thứ 6', 5: 'Thứ 7', 6: 'Chủ Nhật'}
    
    today_name = weekday_map.get(weekday_index, 'Thứ 2')
    menu_today = MENU_STORAGE.get(today_name) # Lấy từ MENU_STORAGE dòng 78
    
    prices = {
        'sang': GIA_SANG,
        'trua': GIA_TRUA,
        'toi': GIA_TOI,
        'tong_ngay': GIA_SANG + GIA_TRUA + GIA_TOI
    }
    
    return menu_today, prices

@app.route('/')
@login_required
def index():
    current_menu, today_name = get_current_menu()
    menu, prices = get_current_menu_with_price()

    # ✅ Tổng học viên
    total_students = StudentStatistic.query.count()

    # 👉 Tổng vắng
    tong_vang_sang = db.session.query(db.func.sum(StudentStatistic.vang_sang)).scalar() or 0
    tong_vang_trua = db.session.query(db.func.sum(StudentStatistic.vang_trua)).scalar() or 0
    tong_vang_toi = db.session.query(db.func.sum(StudentStatistic.vang_toi)).scalar() or 0

    hv_an_sang = total_students - tong_vang_sang
    hv_an_trua = total_students - tong_vang_trua
    hv_an_toi = total_students - tong_vang_toi


    # ===============================
    # 🔥 CBGV (THEO KHOA / ĐƠN VỊ)
    # ===============================

    total_staff = db.session.query(db.func.sum(StaffStatistic.so_luong)).scalar() or 0

    staff_vang_sang = db.session.query(db.func.sum(StaffStatistic.vang_sang)).scalar() or 0
    staff_vang_trua = db.session.query(db.func.sum(StaffStatistic.vang_trua)).scalar() or 0
    staff_vang_toi = db.session.query(db.func.sum(StaffStatistic.vang_toi)).scalar() or 0

    gv_an_sang = total_staff - staff_vang_sang
    gv_an_trua = total_staff - staff_vang_trua
    gv_an_toi = total_staff - staff_vang_toi


    # 👉 GỘP
    stats = {
        'sang': {'hv': hv_an_sang, 'gv': gv_an_sang},
        'trua': {'hv': hv_an_trua, 'gv': gv_an_trua},
        'toi': {'hv': hv_an_toi, 'gv': gv_an_toi}
    }

    return render_template(
        'index.html',
        stats=stats,
        datetime=datetime,
        menu=current_menu,
        today_name=today_name,
        prices=prices,
        total_students=total_students   # 🔥 TRUYỀN SANG HTML
    )

@app.route('/update-meal', methods=['POST'])
@login_required
def update_meal():
    # --- BƯỚC 1: KIỂM TRA QUYỀN (CHẶN ADMIN & HOCVIEN) ---
    if current_user.role in ['admin', 'hocvien']:
        role_name = "Học viên" if current_user.role == 'hocvien' else "Quản trị viên"
        flash(f"Tài khoản {role_name} không được phép đăng ký ăn!", "danger")
        return redirect(url_for('meal_schedule'))

    # --- BƯỚC 2: CHỈ XỬ LÝ NẾU LÀ CBGV ---
    if current_user.role == 'cbgv':
        date_str = request.form.get('date')
        sang = request.form.get('sang') == 'on'
        trua = request.form.get('trua') == 'on'
        toi = request.form.get('toi') == 'on'

        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
            meal = MealRegistration.query.filter_by(
                user_id=current_user.id, 
                date=date_obj
            ).first()

            if not meal:
                meal = MealRegistration(user_id=current_user.id, date=date_obj)
                db.session.add(meal)

            meal.sang = sang
            meal.trua = trua
            meal.toi = toi
            db.session.commit()
            # Trong hàm update_meal của app.py
            new_notification = AdminLog(
            content=f"Đồng chí {current_user.username} đã đăng ký ăn ngày {date_obj}",
            is_read=False # Đánh dấu là chưa đọc để hiện Badge đỏ
        )
            db.session.add(new_notification)
            db.session.commit()    
        
            # --- PHẦN THÊM MỚI: GỬI THÔNG BÁO CHO ADMIN ---
            status_text = []
            if sang: status_text.append("Sáng")
            if trua: status_text.append("Trưa")
            if toi: status_text.append("Tối")
            meals_booked = ", ".join(status_text) if status_text else "Cắt cơm cả ngày"

            new_log = AdminLog(
                content=f"CBGV {current_user.username} đã đăng ký ăn ngày {date_obj.strftime('%d/%m')}: {meals_booked}"
            )
            db.session.add(new_log)
            db.session.commit()
            # ----------------------------------------------
            flash(f"Đã cập nhật đăng ký ăn ngày {date_obj.strftime('%d/%m')} thành công!", "success")
        except Exception as e:
            db.session.rollback()
            flash("Lỗi hệ thống khi lưu dữ liệu!", "danger")
    
    return redirect(url_for('meal_schedule'))

@app.route('/meal-schedule')
@login_required
def meal_schedule():
    today = datetime.now().date()
    start_of_week = today - timedelta(days=today.weekday())

    week_dates = [start_of_week + timedelta(days=i) for i in range(7)]

    meals = MealRegistration.query.filter(
        MealRegistration.user_id == current_user.id,
        MealRegistration.date >= week_dates[0],
        MealRegistration.date <= week_dates[-1]
    ).all()

    meal_dict = {m.date: m for m in meals}

    return render_template(
        'meal_schedule.html',
        week_dates=week_dates,
        meal_dict=meal_dict
    )




# 1. Chức năng chỉnh sửa thực đơn (Admin)
@app.route('/update-menu', methods=['POST'])
@login_required
def update_menu():
    if current_user.role != 'admin':
        return "Không có quyền", 403
    
    day = request.form.get('day')
    sang = request.form.get('sang')
    trua = request.form.get('trua')
    toi = request.form.get('toi')
    
    if day in MENU_STORAGE:
        MENU_STORAGE[day] = {'sang': sang, 'trua': trua, 'toi': toi}
        # Trong thực tế bạn nên lưu vào Database, ở đây ta cập nhật tạm vào biến toàn cục
    return redirect(url_for('index'))

# 2. Chức năng cắt cơm (CBGV và Học viên)
@app.route('/bao-cat-com', methods=['POST'])
@login_required
def bao_cat_com():
    loai_cat = request.form.get('loai_cat') # 'ngay_le', 'thu_7_cn', 'dot_xuat'
    ghi_chu = request.form.get('ghi_chu')
    
    # Logic lưu thông tin cắt cơm vào database tại đây
    # flash(f"Đã gửi yêu cầu cắt cơm: {loai_cat}", "success")
    flash("Đã gửi yêu cầu báo cắt cơm thành công!", "success")
    return redirect(url_for('index'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form['username']).first()
        if user and check_password_hash(user.password, request.form['password']):
            login_user(user)
            # 🔥 LẤY IP
            ip = request.remote_addr
            # 🔥 TẠO LOG
            log = LoginHistory(
                user_id=user.id,
                username=user.username,
                login_time=datetime.now(),
                ip_address=ip
            )
            db.session.add(log)
            db.session.commit()
            # lưu lại log_id để logout cập nhật
            from flask import session
            session['log_id'] = log.id
            # Ghi lại lịch sử đăng nhập
            login_history = LoginHistory(
                user_id=user.id,
                username=user.username,
                login_time=datetime.now(),
                ip_address=ip
            )
            db.session.add(login_history)
            db.session.commit()

            return redirect(url_for('index'))
        flash('Sai tài khoản hoặc mật khẩu!', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    from flask import session

    log_id = session.get('log_id')

    if log_id:
        log = LoginHistory.query.get(log_id)
        if log:
            log.logout_time = datetime.now()
            db.session.commit()

    logout_user()
    return redirect(url_for('login'))

@app.route('/admin/log')
@login_required
def admin_logs():
    # 🔥 CHỈ ADMIN ĐƯỢC XEM
    if current_user.role != 'admin':
        return "❌ Bạn không có quyền truy cập!", 403
    if current_user.role != 'admin':
        return "Không có quyền", 403
    
    logs = AdminLog.query.order_by(AdminLog.created_at.desc()).all()
    # Đánh dấu đã đọc khi admin truy cập
    for log in logs:
        log.is_read = True
    db.session.commit()
    danh_sach_logs = AdminLog.query.order_by(AdminLog.created_at.desc()).all()
    logs = LoginHistory.query.order_by(LoginHistory.login_time.desc()).all()
    return render_template('admin_log.html', logs=danh_sach_logs)

# Route dành cho CBGV: Đăng ký nhanh trưa cả tuần
@app.route('/cbgv/quick-register', methods=['POST'])
@login_required
def quick_register_cbgv():
    if current_user.role != 'cbgv': return "Không có quyền", 403
    today = datetime.now().date()
    start_of_week = today - timedelta(days=today.weekday())
    for i in range(5):
        d = start_of_week + timedelta(days=i)
        reg = MealRegistration.query.filter_by(user_id=current_user.id, date=d).first()
        if not reg:
            reg = MealRegistration(user_id=current_user.id, date=d, trua=True)
            db.session.add(reg)
    db.session.commit()
    flash("Đã đăng ký ăn trưa từ T2-T6!")
    return redirect(url_for('index'))

# --- KHỞI TẠO DB VÀ ADMIN ---
def setup_db():
    with app.app_context():
        db.create_all()
        if not User.query.filter_by(username='admin').first():
            admin = User(
                username='admin', 
                password=generate_password_hash('admin123'),
                role='admin',
                fullname='Quản trị viên'
            )
            db.session.add(admin)
#khởi tạo tài khoản CBGV
        if not User.query.filter_by(username='cbgv').first():
            giaovien = User(
                username='cbgv', 
                password=generate_password_hash('cbgv123'),
                role='cbgv',
                fullname='Cán bộ giáo viên'
            )
            db.session.add(giaovien)
#khởi tạo tài khoản học viên
        if not User.query.filter_by(username='hocvien').first():
            hocvien = User(
                username='hocvien', 
                password=generate_password_hash('hocvien123'),
                role='hocvien',
                fullname='Học viên'
            )
            db.session.add(hocvien)
        db.session.commit()


@app.route('/thong_ke_hoc_vien', methods=['GET', 'POST'])
@login_required
def thong_ke_hoc_vien():
   
    # --- Xử lý thêm lớp ---
    if request.method == 'POST':
         # 🔐 CHẶN KHÔNG PHẢI ADMIN
        if current_user.role != 'admin':
            flash("Bạn không có quyền thực hiện thao tác này!", "danger")
            return redirect(url_for('thong_ke_hoc_vien'))

         # 👉 CHỈ ADMIN MỚI XUỐNG ĐÂY
        action = request.form.get('action')
        if action == 'add_class':
            ten_lop = request.form.get('ten_lop')
  
            if ten_lop:
                ten_lop = ten_lop.strip()

                # 🔍 CHECK TRÙNG
                existing = ClassModel.query.filter(
                    db.func.lower(ClassModel.ten_lop) == ten_lop.lower()
                ).first()

                if existing:
                    flash("Lớp này đã tồn tại!", "danger")
                else:
                    new_class = ClassModel(ten_lop=ten_lop)
                    db.session.add(new_class)
                    db.session.commit()
                    flash("Thêm lớp thành công!", "success")        
        elif action == 'delete_class':
            class_id = request.form.get('class_id')
            cls = ClassModel.query.get(class_id)
            if cls:
                db.session.delete(cls)
                db.session.commit()


        elif action == 'add_student':
            ten_hv = request.form.get('ten_hv')
            lop_id = request.form.get('lop_id')
            if ten_hv and lop_id:
                new_hv = StudentStatistic(ten=ten_hv, lop_id=lop_id)
                db.session.add(new_hv)
                db.session.commit()
        elif action == 'delete_student':
            hv_id = request.form.get('hv_id')
            hv = StudentStatistic.query.get(hv_id)
            if hv:
                db.session.delete(hv)
                db.session.commit()
        elif action == 'update_student':
            hv_id = request.form.get('hv_id')
            hv = StudentStatistic.query.get(hv_id)
            if hv:
                hv.vang_sang = int(request.form.get('vang_sang', 0))
                hv.vang_trua = int(request.form.get('vang_trua', 0))
                hv.vang_toi = int(request.form.get('vang_toi', 0))
                hv.tien_thua = hv.vang_sang * 15000 + hv.vang_trua * 30000 + hv.vang_toi * 30000
                db.session.commit()

    # --- Lấy dữ liệu để hiển thị ---
    lop_filter = request.args.get('lop_id', type=int)
    classes = ClassModel.query.all()
    danh_sach_hv = []

    if lop_filter:
        danh_sach_hv = StudentStatistic.query.filter_by(lop_id=lop_filter).all()
    elif classes:
        # Lớp đầu tiên
        danh_sach_hv = StudentStatistic.query.filter_by(lop_id=classes[0].id).all()
        lop_filter = classes[0].id

    return render_template(
        'thong_ke_hoc_vien.html',
        classes=classes,
        danh_sach_hv=danh_sach_hv,
        lop_id=int(lop_filter) if lop_filter else None
    )

    
    
    
# Danh sách giả lập để lưu phản hồi (trong thực tế sẽ lưu vào Database)
feedbacks = [
    {'id': 1, 'nguoi_gui': 'Nguyễn Văn A', 'noi_dung': 'Cơm hôm nay hơi khô', 'ngay': '30/03/2026', 'trang_thai': 'Chờ xử lý'},
    {'id': 2, 'nguoi_gui': 'Trần Thị B', 'noi_dung': 'Món cá kho rất ngon', 'ngay': '29/03/2026', 'trang_thai': 'Đã tiếp nhận'}
]

@app.route('/y-kien-phan-hoi', methods=['GET', 'POST'])
def y_kien_phan_hoi():
    if request.method == 'POST':
        # 1. Lấy dữ liệu từ ô nhập liệu trong form
        noi_dung_moi = request.form.get('noi_dung')
        
        if noi_dung_moi:
            # 2. Tạo một bản ghi phản hồi mới
            new_entry = {
                'nguoi_gui': 'Học viên ẩn danh', # Sau này bạn có thể thay bằng tên user đăng nhập
                'noi_dung': noi_dung_moi,
                'ngay': datetime.now().strftime("%d/%m/%Y"),
                'trang_thai': 'Chờ xử lý'
            }
            # 3. Thêm vào đầu danh sách để nó hiện lên trên cùng
            msg = Message(
                sender=current_user.username if current_user.is_authenticated else "Ẩn danh",
                role=current_user.role if current_user.is_authenticated else "guest",
                content=noi_dung_moi
            )

            db.session.add(msg)
            db.session.commit()
            
        # 4. Sau khi xử lý xong, chuyển hướng lại chính trang này để cập nhật bảng
        return redirect(url_for('y_kien_phan_hoi'))
    
    # Nếu là yêu cầu GET (vào trang bình thường) thì hiện danh sách feedbacks
    return render_template('y_kien.html', feedbacks=feedbacks)

@app.context_processor
def inject_now():
    # Hàm này giúp biến 'thoi_gian_hien_tai' có thể được gọi ở bất kỳ file HTML nào
    return {'thoi_gian_hien_tai': datetime.now()}

@app.route('/send-admin-message', methods=['POST'])
@login_required
def send_admin_message():
    data = request.get_json()
    noi_dung = data.get('message')

    msg = Message(
        sender=current_user.username,
        role=current_user.role,
        content=noi_dung
    )

    db.session.add(msg)
    db.session.commit()

    return jsonify(success=True)

@app.route('/get-messages')
@login_required
def get_messages():
    messages = Message.query.order_by(Message.created_at).all()

    return jsonify([
        {
            "id": m.id,
            "sender": m.sender,
            "role": m.role,
            "content": m.content,
            "reply_to": m.reply_to
        } for m in messages
    ])

@app.route('/admin/messages')
@login_required
def admin_messages():
    if current_user.role != 'admin':
        return "Không có quyền", 403

    messages = Message.query.order_by(Message.created_at.desc()).all()
    return render_template('admin_messages.html', messages=messages)

@app.route('/admin/reply/<int:msg_id>', methods=['POST'])
@login_required
def reply_message(msg_id):
    if current_user.role != 'admin':
        return "Không có quyền", 403

    content = request.form.get('reply')

    reply = Message(
        sender=current_user.username,
        role='admin',
        content=content,
        reply_to=msg_id
    )

    db.session.add(reply)
    db.session.commit()

    return redirect(url_for('admin_messages'))



@app.route('/update_stats_bulk', methods=['POST'])
@login_required
def update_stats_bulk():
    if current_user.role != 'admin':
        flash("Bạn không có quyền thực hiện thao tác này!", "danger")
        return redirect(url_for('thong_ke_hoc_vien'))
    # 🔒 Chỉ admin
    if current_user.role != 'admin':
        return "Không có quyền", 403

    lop_id = request.form.get('lop_id')

    try:
        # 👉 LẤY ĐÚNG BẢNG
        danh_sach_hv = StudentStatistic.query.filter_by(lop_id=lop_id).all()

        for hv in danh_sach_hv:
            # 👉 LẤY DATA TỪ INPUT
            v_sang = int(request.form.get(f'vang_sang_{hv.id}', 0))
            v_trua = int(request.form.get(f'vang_trua_{hv.id}', 0))
            v_toi = int(request.form.get(f'vang_toi_{hv.id}', 0))

            # 👉 UPDATE MODEL
            hv.vang_sang = v_sang
            hv.vang_trua = v_trua
            hv.vang_toi = v_toi

            # 👉 TÍNH TIỀN
            hv.tien_thua = (v_sang * 15000) + (v_trua * 30000) + (v_toi * 30000)

        db.session.commit()
        flash("Lưu thành công!", "success")

    except Exception as e:
        db.session.rollback()
        print("LỖI:", e)
        flash("Lỗi khi lưu!", "danger")

    return redirect(url_for('thong_ke_hoc_vien', lop_id=lop_id))

@app.route('/export_excel')
@login_required
def export_excel():
    if current_user.role != 'admin':
        return "Không có quyền", 403

    wb = Workbook()
    wb.remove(wb.active)  # Xóa sheet mặc định

    # ===== STYLE =====
    bold_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=16)
    center_align = Alignment(horizontal='center', vertical='center')

    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===== LẤY TẤT CẢ LỚP =====
    classes = ClassModel.query.all()

    for cls in classes:
        ws = wb.create_sheet(title=cls.ten_lop)

        # ===== TIÊU ĐỀ =====
        ws.merge_cells('A1:E1')
        ws['A1'] = f"THỐNG KÊ HỌC VIÊN - {cls.ten_lop}"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align

        # ===== HEADER =====
        headers = ["Tên", "Vắng sáng", "Vắng trưa", "Vắng tối", "Tiền thừa"]
        ws.append(headers)

        for col in range(1, 6):
            cell = ws.cell(row=2, column=col)
            cell.font = bold_font
            cell.alignment = center_align
            cell.fill = header_fill
            cell.border = thin_border

        # ===== DATA =====
        students = StudentStatistic.query.filter_by(lop_id=cls.id).all()

        tong = 0
        row_index = 3

        for hv in students:
            ws.cell(row=row_index, column=1, value=hv.ten)
            ws.cell(row=row_index, column=2, value=hv.vang_sang)
            ws.cell(row=row_index, column=3, value=hv.vang_trua)
            ws.cell(row=row_index, column=4, value=hv.vang_toi)
            ws.cell(row=row_index, column=5, value=hv.tien_thua)

            for col in range(1, 6):
                cell = ws.cell(row=row_index, column=col)
                cell.border = thin_border
                if col != 1:
                    cell.alignment = center_align

            tong += hv.tien_thua
            row_index += 1

        # ===== TỔNG =====
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=4)
        ws.cell(row=row_index, column=1, value="TỔNG").font = bold_font
        ws.cell(row=row_index, column=1).alignment = center_align

        tong_cell = ws.cell(row=row_index, column=5, value=tong)
        tong_cell.font = bold_font
        tong_cell.border = thin_border

        # ===== FORMAT TIỀN =====
        for row in ws.iter_rows(min_row=3, min_col=5, max_col=5, max_row=row_index):
            for cell in row:
                cell.number_format = '#,##0'

        # ===== AUTO WIDTH =====
        column_widths = [25, 15, 15, 15, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

    # ===== XUẤT FILE =====
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="thong_ke_tat_ca_lop.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
@app.context_processor
def inject_globals():
    # Khởi tạo giá trị mặc định là None
    admin_log_cls = None
    message_cls = None
    
    try:
        # Thử import nếu đồng chí có file models.py riêng
        from models import AdminLog, Message
        admin_log_cls = AdminLog
        message_cls = Message
    except (ImportError, ModuleNotFoundError):
        # Nếu không thấy file models.py, thử lấy trực tiếp trong app.py
        try:
            from __main__ import AdminLog, Message
            admin_log_cls = AdminLog
            message_cls = Message
        except ImportError:
            # Nếu vẫn không thấy, tìm trong scope hiện tại
            import sys
            current_module = sys.modules[__name__]
            admin_log_cls = getattr(current_module, 'AdminLog', None)
            message_cls = getattr(current_module, 'Message', None)

    return dict(AdminLog=admin_log_cls, Message=message_cls)
    


if __name__ == '__main__':
    setup_db()
    app.run(debug=True, host='0.0.0.0', port=5000)